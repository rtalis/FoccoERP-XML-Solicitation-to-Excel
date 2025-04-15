
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import sys
from datetime import datetime
import os
import re
import unicodedata
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Get the path of the script

bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
path_to_model1 = os.path.abspath(os.path.join(bundle_dir, 'model1.xlsx'))
path_to_model2 = os.path.abspath(os.path.join(bundle_dir, 'model2.xlsx'))

def get_text(element, tag):
    child = element.find(tag)
    return child.text if child is not None else ''

def write_to_excel(sheet, row, data):
    columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, value in zip(columns, data):
        sheet[f"{col}{row}"] = value
        
def remove_diacritics(text):
    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )

def extract_solicitacao(text):
    text = remove_diacritics(text)
    match = re.search(r'SOLICITACAO?\s*(\d+)', text, re.IGNORECASE)
    return match.group(1) if match else ''

def purchase_order_to_excel(excel_file, xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        row = 3
        for item in root.findall(".//TPED_COMPRA"):
            observacao = get_text(item, "OBSERVACAO")
            solicitacao_num = "" if observacao is None else extract_solicitacao(observacao)
            data = [
                "", 
                solicitacao_num,# Column B
                get_text(item, "COD_PEDC"),  
                get_text(item, "FOR_DESCRICAO"),  
                get_text(item, "DT_EMIS"),  
                float(get_text(item, "TOT_LIQUIDO_IPI1").replace(',', '.') or 0.0),  # Column F
                "", ""  # Columns G, H
            ]
            write_to_excel(sheet, row, data)
            row += 1

        new_filename = f'pedidos de compra {datetime.now().strftime("%Y-%m-%d %H-%M-%S")}.xlsx'
        workbook.save(new_filename)
        print(f"Saved file as {new_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")
        input("Aperte enter para sair")

def solicitaton_to_excel(excel_file, xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        row = 3
        for item in root.findall('.//G_ITENS_SOLICITACAO'):
            data = [
                "",  # Column A
                get_text(item, 'DESCRICAO'),  # Column B
                get_text(item, 'QTDE'),  # Column C
                "", "", "", "", ""  # Columns D, E, F, G, H
            ]
            write_to_excel(sheet, row, data)
            row += 1

        new_filename = f'cotação {datetime.now().strftime("%Y-%m-%d %H-%M-%S")}.xlsx'
        workbook.save(new_filename)
        print(f"Saved file as {new_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")
        input("Aperte enter para sair")
        
        
def cotacao_rcot0300_to_excel(excel_file, xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        row = 3
        for g4_item in root.findall(".//G_4"):
            data = [
                "",  # Column A
                get_text(g4_item, "DESC_ITEM"),  # Column B
                get_text(g4_item, "QTDE"),       # Column C
                "", "", "", "", ""               # Columns D through H
            ]
            write_to_excel(sheet, row, data)
            row += 1

        new_filename = f'cotação RCOT0300 {datetime.now().strftime("%Y-%m-%d %H-%M-%S")}.xlsx'
        workbook.save(new_filename)
        print(f"Saved file as {new_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")
        input("Aperte enter para sair")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        root = Tk()
        root.withdraw()
        xml_file = askopenfilename(filetypes=[("XML files", "*.xml")])
        if not xml_file:
            print("No file selected. Exiting.")
            sys.exit(1)
    else:
        xml_file = sys.argv[1]

    print(f"Processing file: {xml_file}")

    file_name = os.path.basename(xml_file)
    if file_name.startswith("RPDC0250_RUAH"):
        purchase_order_to_excel(path_to_model2, xml_file)
    elif file_name.startswith("RPDC0251"):
        solicitaton_to_excel(path_to_model1, xml_file)
    elif file_name.startswith("RCOT0300"):
        cotacao_rcot0300_to_excel(path_to_model1, xml_file)